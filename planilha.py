from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment

def criaPlan(nomeArq, produtos):
    #Cria a planilha e suas primeiras linhas de titulos.
    wb = Workbook()
    plan =wb.active
    plan.title = nomeArq
    plan['A1']=plan.title.capitalize()
    plan['A2']='Nome do Produto'
    plan['B2']='Valor'

    plan.merge_cells('A1:B1')
    plan['A1'].alignment = Alignment(horizontal='center', vertical='center')
    celulas = ['A1','A2','B2']
    estiliza(plan,celulas,1)


    addProdutos(produtos, plan, 3)
    wb.save(f'{nomeArq}.xlsx')

def addProdutos(produtos, plan, linhaInicial):
    #Adiciona as pessoas e seus dados na planilha.
    linha = linhaInicial
    for p in produtos:
        plan[f'A{linha}']=p['nome']
        plan[f'B{linha}']=p['valor']

        celulas = [f'A{linha}',f'B{linha}']
        estiliza(plan,celulas,2)
        linha +=1

def estiliza(planilha,celulas,estilo):
    # Adiciona estilo às celulas recebidas em lista.
    # Caso estilo for 1, a celula assumirá estilo de título com background escuro e fontes claras.
    # Caso estilo for 2, a celula assumirá estilo de dados com background claro e fontes escuras.
    if estilo == 1:
        cor = 'ffffff'
        corFun = '00008B'
        b = True
    elif estilo ==2:
        cor ='000000'
        corFun = '87CEFA'
        b = False
    else: return 0
    for i in celulas:
        planilha[i].font = Font(bold=b,color=cor)
        planilha[i].fill = PatternFill('solid', fgColor=corFun)
        larg = Side(border_style='thin',color='000000')
        planilha[i].border = Border(top=larg,bottom=larg,right=larg,left=larg)